"""Document parsers: PDF, DOCX, XLSX, plain text."""

from __future__ import annotations

from pathlib import Path


def parse_pdf(path: Path) -> str:
    """Extract text from PDF using pymupdf (fitz)."""
    try:
        import fitz  # pymupdf
    except ImportError:
        raise ImportError("pip install pymupdf  (required for PDF ingestion)")
    with fitz.open(str(path)) as doc:
        return "\n\n".join(page.get_text() for page in doc)


def parse_docx(path: Path) -> str:
    """Extract text from DOCX using python-docx."""
    try:
        from docx import Document
    except ImportError:
        raise ImportError("pip install python-docx  (required for DOCX ingestion)")
    doc = Document(str(path))
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def parse_xlsx(path: Path) -> str:
    """Extract text from XLSX using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("pip install openpyxl  (required for XLSX ingestion)")
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        parts: list[str] = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                parts.append(f"## {sheet}\n" + "\n".join(rows))
        return "\n\n".join(parts)
    finally:
        wb.close()


def parse_text(path: Path) -> str:
    """Read plain text file."""
    return path.read_text(encoding="utf-8", errors="replace")


from typing import Callable

DISPATCH: dict[str, Callable] = {
    ".pdf": parse_pdf,
    ".docx": parse_docx,
    ".xlsx": parse_xlsx,
    ".xls": parse_xlsx,
    ".txt": parse_text,
    ".md": parse_text,
    ".csv": parse_text,
}


def parse_document(path: Path) -> str:
    """Auto-dispatch based on file extension."""
    ext = path.suffix.lower()
    parser = DISPATCH.get(ext)
    if not parser:
        raise ValueError(f"Unsupported file type: {ext}")
    return parser(path)
